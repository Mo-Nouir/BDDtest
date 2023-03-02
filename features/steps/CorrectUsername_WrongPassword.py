from behave import *
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from test_help.TestPage import locators
from test_help.GetData import write_to_csv


@when(u'I enter a correct username and a wrong password')
def step_impl(context):
    context.locators = locators()
    workbook = openpyxl.load_workbook("test_help/DataDrivenTesting.xlsx")
    sheet = workbook.get_sheet_by_name("Sheet1")
    # get the Username value from excel
    Input_Username = sheet.cell(row=5, column=1).value
    Input_Password = sheet.cell(row=5, column=2).value
    print(Input_Password, Input_Username)
    email = context.driver.find_element(By.XPATH, locators.email)
    email.send_keys(Input_Username)
    password = context.driver.find_element(By.XPATH, locators.password)
    password.send_keys(Input_Password)


@then(u'the text "Your password is invalid!" is displayed')
def step_impl(context):
    email_error = context.driver.find_element(By.XPATH, locators.password_error).text
    write_to_csv([[email_error]])
    assert email_error == "Your password is invalid!"

