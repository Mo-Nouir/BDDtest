import csv
import random
import string
import openpyxl

def write_to_csv(data):
    workbook = openpyxl.load_workbook("test_help/DataDrivenTesting.xlsx")
    sheet = workbook.get_sheet_by_name("Sheet1")
    # get the Username value from excel
    Input_Username = sheet.cell(row=2, column=1).value
    Input_Password = sheet.cell(row=2, column=2).value



def random_username(length):
    letters = string.ascii_letters
    return ''.join(random.choice(letters) for i in range(length))


def random_password(length):
    chars = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(chars) for i in range(length))
