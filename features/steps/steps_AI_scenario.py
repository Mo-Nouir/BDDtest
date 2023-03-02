import time
import openpyxl
from behave import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from test_help.TestPage import locators
import undetected_chromedriver as uc
import pandas as pd
from test_help.GetData import random_password, random_username


@when(u'I enter a random username and a random password')
def step_impl(context):
    Input_Username = random_username(7)
    Input_Password = random_password(10)
    print(Input_Password, Input_Username)
    # Load the existing Excel file
    wb = openpyxl.load_workbook("test_help/DataDrivenTesting.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    sheet = wb.active

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Add the values to the new row
    sheet.cell(row=next_row, column=1).value = Input_Username
    sheet.cell(row=next_row, column=2).value = Input_Password

    # Save the changes to the Excel file
    wb.save("test_help/DataDrivenTesting.xlsx")

    email = context.driver.find_element(By.XPATH, locators.email)
    email.send_keys(Input_Username)
    password = context.driver.find_element(By.XPATH, locators.password)
    password.send_keys(Input_Password)


@then(u'Store the shown message in the csv file')
def step_impl(context):

    wb = openpyxl.load_workbook("test_help/DataDrivenTesting.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    sheet = wb.active
    if context.driver.current_url == "https://practicetestautomation.com/logged-in-successfully/":
        sheet.cell(row=sheet.max_row, column=3).value = "successfully logged in"
        wb.save("test_help/DataDrivenTesting.xlsx")
    else:
        lol = context.driver.find_element(By.XPATH, locators.password_error).text
        print(lol)
        sheet.cell(row=sheet.max_row, column=3).value = lol
        wb.save("test_help/DataDrivenTesting.xlsx")
