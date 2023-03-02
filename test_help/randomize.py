import openpyxl
import random
import string

def random_username(length):
    letters = string.ascii_letters
    return ''.join(random.choice(letters) for i in range(length))

def random_password(length):
    chars = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(chars) for i in range(length))


# Create a new Excel file and add a header row
workbook = openpyxl.load_workbook("DataDrivenTesting.xlsx")
sheet = workbook.get_sheet_by_name("Sheet1")


# Generate random user names and passwords and add them to the sheet
for i in range(10):
    username = random_username(10)
    password = random_password(15)
    sheet.cell(row=i+7, column=1).value = username
    sheet.cell(row=i+7, column=2).value = password

# Save the Excel file
sheet.save("DataDrivenTesting.xlsx")