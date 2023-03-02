import time
import openpyxl
from behave import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from test_help.TestPage import locators
import undetected_chromedriver as uc
import csv
from test_help.GetData import write_to_csv

@given(u'I open the Chrome browser in incognito mode')
def step_impl(context):
    context.locators=locators()
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    context.driver = webdriver.Chrome(options=chrome_options, executable_path="C:\chromedriver.exe")


@given(u'I navigate to "https://practicetestautomation.com/practice-test-login/"')
def step_impl(context):
    context.driver.get(locators.testURL)


@when(u'I enter an invalid username and a valid password')
def step_impl(context):
    workbook = openpyxl.load_workbook("test_help/DataDrivenTesting.xlsx")
    sheet = workbook.get_sheet_by_name("Sheet1")
    # get the Username value from excel
    Input_Username = sheet.cell(row=2, column=1).value
    Input_Password = sheet.cell(row=2, column=2).value
    print(Input_Password, Input_Username)
    email = context.driver.find_element(By.XPATH, locators.email)
    email.send_keys(Input_Username)
    password = context.driver.find_element(By.XPATH, locators.password)
    password.send_keys(Input_Password)


@when(u'I click on the "Submit" button')
def step_impl(context):
    context.driver.find_element(By.XPATH, locators.button).click()


@then(u'the error text should be displayed')
def step_impl(context):
    time.sleep(2)
    error_box = context.driver.find_element(By.XPATH, locators.password_error).is_displayed
    assert error_box


@then(u'the text "Your username is invalid!" is displayed')
def step_impl(context):
    password_error = context.driver.find_element(By.XPATH, locators.email_error).text
    write_to_csv([[password_error]])
    assert password_error == "Your username is invalid!"
