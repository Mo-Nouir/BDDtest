import time
import openpyxl
from behave import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from test_help.TestPage import locators
#from test_help.GetData import write_to_csv


@when(u'I enter a correct email address and a correct password')
def step_impl(context):
    context.locators=locators()
    workbook = openpyxl.load_workbook("test_help/DataDrivenTesting.xlsx")
    sheet = workbook.get_sheet_by_name("Sheet1")
    # get the Username value from excel
    Input_Username = sheet.cell(row=6, column=1).value
    Input_Password = sheet.cell(row=6, column=2).value
    print(Input_Password, Input_Username)
    email = context.driver.find_element(By.XPATH, locators.email)
    email.send_keys(Input_Username)
    password = context.driver.find_element(By.XPATH, locators.password)
    password.send_keys(Input_Password)


@then(u'logged in successfully')
def step_impl(context):
    time.sleep(5)
    login = context.driver.find_element(By.XPATH, locators.LogOutButton).is_displayed()
    #write_to_csv([["successfully logged in"]])
    assert login
